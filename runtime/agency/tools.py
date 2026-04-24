"""Tools the agent can call. Designed to be safe by default.

Each tool has:
 - a JSON-schema definition (fed to Claude)
 - a Python implementation that returns a string result

Shell access is gated by AGENCY_ALLOW_SHELL (default off). When enabled, only
commands whose first token is in SHELL_ALLOWLIST run. Network fetch is a
read-only GET over http/https with private-range hosts blocked (see
`_is_private_host`).
"""

from __future__ import annotations

import os
import shlex
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import httpx

DEFAULT_SHELL_ALLOWLIST = (
    "ls", "cat", "head", "tail", "wc", "find", "grep", "rg",
    "pwd", "echo", "date", "git", "tree", "stat", "file",
    "python", "python3", "node", "npm", "pip", "uv",
    "make", "ruff", "mypy", "pytest",
)

MAX_OUTPUT_BYTES = 64_000


@dataclass
class ToolResult:
    content: str
    is_error: bool = False


@dataclass
class Tool:
    name: str
    description: str
    input_schema: dict[str, Any]
    func: Callable[[dict[str, Any], "ToolContext"], ToolResult]

    def to_anthropic(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "input_schema": self.input_schema,
        }


@dataclass
class ToolContext:
    """Per-session execution context passed to every tool call."""

    workdir: Path
    allow_shell: bool = False
    shell_allowlist: tuple[str, ...] = DEFAULT_SHELL_ALLOWLIST
    allow_network: bool = True
    allow_computer_use: bool = False
    timeout_s: int = 30

    @classmethod
    def from_env(cls, workdir: Path | None = None) -> "ToolContext":
        try:
            timeout_s = int(os.environ.get("AGENCY_TOOL_TIMEOUT", "30"))
        except ValueError:
            timeout_s = 30
        return cls(
            workdir=(workdir or Path.cwd()).resolve(),
            allow_shell=_truthy(os.environ.get("AGENCY_ALLOW_SHELL")),
            allow_network=not _truthy(os.environ.get("AGENCY_NO_NETWORK")),
            allow_computer_use=_truthy(os.environ.get("AGENCY_ENABLE_COMPUTER_USE")),
            timeout_s=timeout_s,
        )


def _truthy(v: str | None) -> bool:
    return (v or "").strip().lower() in ("1", "true", "yes", "on")


def _safe_path(ctx: ToolContext, raw: str) -> Path:
    """Resolve `raw` against the workdir and reject anything escaping it."""
    candidate = (ctx.workdir / raw).resolve() if not os.path.isabs(raw) else Path(raw).resolve()
    try:
        candidate.relative_to(ctx.workdir)
    except ValueError as e:
        raise PermissionError(f"Path escapes workdir: {raw}") from e
    return candidate


def _truncate(text: str) -> str:
    """Return `text` trimmed to at most MAX_OUTPUT_BYTES encoded bytes.

    Does a binary search on a character prefix so the final encoded size —
    including the truncation marker — fits the cap even for multi-byte UTF-8.
    """
    encoded_len = len(text.encode("utf-8", errors="ignore"))
    if encoded_len <= MAX_OUTPUT_BYTES:
        return text

    low, high = 0, len(text)
    best = ""
    best_suffix = ""
    while low <= high:
        mid = (low + high) // 2
        cut = text[:mid]
        suffix = f"\n\n[... truncated, {len(text) - mid} chars omitted]"
        total = len(cut.encode("utf-8")) + len(suffix.encode("utf-8"))
        if total <= MAX_OUTPUT_BYTES:
            best, best_suffix = cut, suffix
            low = mid + 1
        else:
            high = mid - 1
    return best + best_suffix


def _is_private_host(host: str) -> bool:
    """Reject hosts that resolve to loopback, link-local, private, or metadata ranges.

    Uses `socket.getaddrinfo` to cover every resolved IP; any private or
    otherwise-unsafe target disqualifies the host.
    """
    import ipaddress
    import socket

    # Quick literal checks on the host string.
    lowered = host.lower().strip()
    if lowered in ("localhost", "metadata", "metadata.google.internal",
                   "instance-data", "metadata.internal"):
        return True

    try:
        infos = socket.getaddrinfo(host, None)
    except (socket.gaierror, UnicodeError):
        # DNS failure: treat as private — safer to reject than leak.
        return True

    for info in infos:
        sockaddr = info[4]
        ip_str = sockaddr[0].split("%")[0]  # strip zone id if present
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            return True
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_multicast or ip.is_reserved
                or ip.is_unspecified):
            return True
        # AWS / GCP metadata endpoints use 169.254.169.254 (caught by link_local
        # above) and fd00:ec2::254 (caught by is_private). Extra safety belt:
        if ip_str == "169.254.169.254":
            return True
    return False


# ----- Tool implementations -----------------------------------------------

def _read_file(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    path = _safe_path(ctx, args["path"])
    if not path.exists():
        return ToolResult(f"File not found: {args['path']}", is_error=True)
    if path.is_dir():
        return ToolResult(f"Path is a directory: {args['path']}", is_error=True)
    try:
        return ToolResult(_truncate(path.read_text(encoding="utf-8", errors="replace")))
    except OSError as e:
        return ToolResult(f"Read error: {e}", is_error=True)


def _list_dir(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    path = _safe_path(ctx, args.get("path", "."))
    if not path.exists():
        return ToolResult(f"Path not found: {args.get('path', '.')}", is_error=True)
    if not path.is_dir():
        return ToolResult(f"Not a directory: {path}", is_error=True)
    entries = []
    for entry in sorted(path.iterdir()):
        kind = "d" if entry.is_dir() else "f"
        entries.append(f"{kind} {entry.name}")
    return ToolResult("\n".join(entries) or "(empty)")


def _write_file(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    path = _safe_path(ctx, args["path"])
    content = args.get("content", "")
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        path.write_text(content, encoding="utf-8")
    except OSError as e:
        return ToolResult(f"Write error: {e}", is_error=True)
    return ToolResult(f"Wrote {len(content)} chars to {path.relative_to(ctx.workdir)}")


def _edit_file(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """Replace an exact string with another in an existing file.

    Modeled on Anthropic's `str_replace_based_edit_tool`. `old_string` must
    appear exactly once; ambiguous or missing matches are errors.
    """
    path = _safe_path(ctx, args["path"])
    old = args["old_string"]
    new = args.get("new_string", "")
    if not path.exists():
        return ToolResult(f"File not found: {args['path']}", is_error=True)
    if path.is_dir():
        return ToolResult(f"Path is a directory: {args['path']}", is_error=True)
    try:
        text = path.read_text(encoding="utf-8")
    except OSError as e:
        return ToolResult(f"Read error: {e}", is_error=True)
    count = text.count(old)
    if count == 0:
        return ToolResult("old_string not found in file.", is_error=True)
    if count > 1:
        return ToolResult(
            f"old_string matches {count} places; must be unique. Add more context around it.",
            is_error=True,
        )
    try:
        path.write_text(text.replace(old, new, 1), encoding="utf-8")
    except OSError as e:
        return ToolResult(f"Write error: {e}", is_error=True)
    rel = path.relative_to(ctx.workdir)
    return ToolResult(f"Replaced 1 occurrence in {rel}.")


def _extract_doc(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """Extract plain text from a document.

    Supports .pdf (pypdf), .docx (python-docx), .xlsx (openpyxl), and falls
    back to UTF-8 read for text-ish formats. Optional deps; missing ones
    surface a clear install hint instead of a stack trace.
    """
    path = _safe_path(ctx, args["path"])
    if not path.exists():
        return ToolResult(f"File not found: {args['path']}", is_error=True)
    if path.is_dir():
        return ToolResult(f"Path is a directory: {args['path']}", is_error=True)

    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return ToolResult(_truncate(_extract_pdf(path)))
        if suffix == ".docx":
            return ToolResult(_truncate(_extract_docx(path)))
        if suffix == ".xlsx":
            return ToolResult(_truncate(_extract_xlsx(path)))
        # Text-ish fallback.
        return ToolResult(_truncate(path.read_text(encoding="utf-8", errors="replace")))
    except _MissingDep as e:
        return ToolResult(str(e), is_error=True)
    except Exception as e:  # noqa: BLE001 - surface to model
        return ToolResult(f"Extract error: {type(e).__name__}: {e}", is_error=True)


class _MissingDep(RuntimeError):
    pass


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore[import-not-found]
    except ImportError as e:
        raise _MissingDep(
            "PDF extraction requires `pypdf`. Install with: pip install -e 'runtime[docs]'"
        ) from e
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages, 1):
        pages.append(f"--- page {i} ---\n{page.extract_text() or ''}")
    return "\n".join(pages) or "(no extractable text)"


def _extract_docx(path: Path) -> str:
    try:
        import docx  # type: ignore[import-not-found]
    except ImportError as e:
        raise _MissingDep(
            "DOCX extraction requires `python-docx`. Install with: pip install -e 'runtime[docs]'"
        ) from e
    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path: Path) -> str:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as e:
        raise _MissingDep(
            "XLSX extraction requires `openpyxl`. Install with: pip install -e 'runtime[docs]'"
        ) from e
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"=== sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            parts.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(parts)


def _run_shell(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    if not ctx.allow_shell:
        return ToolResult(
            "Shell access is disabled. Set AGENCY_ALLOW_SHELL=1 to enable "
            "(commands are still gated by an allowlist).",
            is_error=True,
        )
    command = args["command"].strip()
    if not command:
        return ToolResult("Empty command.", is_error=True)
    try:
        tokens = shlex.split(command)
    except ValueError as e:
        return ToolResult(f"Could not parse command: {e}", is_error=True)
    head = tokens[0]
    if head not in ctx.shell_allowlist:
        return ToolResult(
            f"Command not in allowlist: {head}. Allowed: {', '.join(sorted(ctx.shell_allowlist))}",
            is_error=True,
        )
    if shutil.which(head) is None:
        return ToolResult(f"Executable not found: {head}", is_error=True)
    try:
        proc = subprocess.run(
            tokens,
            cwd=str(ctx.workdir),
            capture_output=True,
            text=True,
            timeout=ctx.timeout_s,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return ToolResult(f"Timed out after {ctx.timeout_s}s.", is_error=True)
    except OSError as e:
        return ToolResult(f"Exec error: {e}", is_error=True)
    out = proc.stdout
    if proc.stderr:
        out = (out + "\n[stderr]\n" + proc.stderr) if out else proc.stderr
    out = out + f"\n[exit: {proc.returncode}]"
    return ToolResult(_truncate(out), is_error=proc.returncode != 0)


def _web_fetch(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """Fetch an http(s) URL with SSRF protection, manual redirect control, and byte cap.

    - Rejects URLs that resolve to private / loopback / link-local / metadata IPs.
    - Re-checks the target of every redirect hop (max 5).
    - Streams the response, stops when the body hits MAX_OUTPUT_BYTES.
    """
    if not ctx.allow_network:
        return ToolResult("Network access disabled (AGENCY_NO_NETWORK).", is_error=True)
    url = args["url"]
    if not (url.startswith("https://") or url.startswith("http://")):
        return ToolResult("Only http/https URLs are allowed.", is_error=True)

    try:
        with httpx.Client(
            follow_redirects=False, timeout=ctx.timeout_s,
            headers={"User-Agent": "agency-runtime/0.1"},
        ) as client:
            for hop in range(6):
                parsed = httpx.URL(url)
                host = parsed.host
                if not host:
                    return ToolResult("URL has no host.", is_error=True)
                if _is_private_host(host):
                    return ToolResult(
                        f"Refusing to fetch {host}: private / loopback / metadata address.",
                        is_error=True,
                    )
                req = client.build_request("GET", url)
                with client.send(req, stream=True) as resp:
                    if 300 <= resp.status_code < 400 and "location" in resp.headers:
                        url = str(resp.next_request.url if resp.next_request else resp.headers["location"])
                        resp.close()
                        continue
                    prefix = f"HTTP {resp.status_code}\n\n"
                    prefix_bytes = prefix.encode("utf-8")
                    remaining = max(MAX_OUTPUT_BYTES - len(prefix_bytes), 0)
                    buf = bytearray()
                    if remaining > 0:
                        for chunk in resp.iter_bytes():
                            if not chunk:
                                continue
                            if len(chunk) >= remaining:
                                buf.extend(chunk[:remaining])
                                break
                            buf.extend(chunk)
                            remaining -= len(chunk)
                    body = buf.decode("utf-8", errors="replace")
                    return ToolResult(prefix + body)
            return ToolResult("Too many redirects (>5).", is_error=True)
    except httpx.HTTPError as e:
        return ToolResult(f"Fetch error: {e}", is_error=True)


def _list_skills(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    # Filled in at registry build time so the executor can introspect siblings.
    skills = getattr(ctx, "_skills_summary", None)
    if not skills:
        return ToolResult("No sibling skills are registered in this session.", is_error=True)
    return ToolResult(skills)


def _delegate_to_skill(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """Invoke another skill as a subagent and return its final text.

    Requires the executor to have populated `ctx._delegate_runner`. Enforces
    a small recursion cap so delegation cycles can't explode.
    """
    runner = getattr(ctx, "_delegate_runner", None)
    if runner is None:
        return ToolResult("Delegation is not available in this session.", is_error=True)
    slug = args.get("slug", "").strip()
    request = args.get("request", "").strip()
    if not slug or not request:
        return ToolResult("delegate_to_skill requires 'slug' and 'request'.", is_error=True)
    try:
        return ToolResult(runner(slug, request))
    except KeyError as e:
        return ToolResult(f"No such skill: {e}", is_error=True)
    except RecursionError as e:
        return ToolResult(str(e), is_error=True)


def _computer_use(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """Dispatch to pyautogui for mouse/keyboard/screenshot actions.

    Implements Anthropic's `computer_20250124` client side. Requires a
    display (X11/Wayland on Linux, native on macOS/Windows) and the
    `[computer]` extra. Disabled unless AGENCY_ENABLE_COMPUTER_USE=1.

    ⚠️ Gives the agent full UI control of the running user session. Only
    enable in a sandbox or dedicated VM.
    """
    if not ctx.allow_computer_use:
        return ToolResult(
            "Computer use is disabled. Set AGENCY_ENABLE_COMPUTER_USE=1 AND run "
            "on a machine with a display + `pip install -e 'runtime[computer]'`.",
            is_error=True,
        )
    try:
        import base64
        import io

        import pyautogui  # type: ignore[import-not-found]
        from PIL import Image  # noqa: F401 — used only to verify install
    except ImportError as e:
        return ToolResult(
            "Missing deps: pip install -e 'runtime[computer]' (pyautogui, pillow).",
            is_error=True,
        )
    except Exception as e:  # noqa: BLE001 — e.g. no DISPLAY
        return ToolResult(
            f"pyautogui init failed ({type(e).__name__}: {e}). Need a display.",
            is_error=True,
        )

    action = (args.get("action") or "").strip()
    coord = args.get("coordinate")
    text = args.get("text")

    try:
        if action == "screenshot":
            img = pyautogui.screenshot()
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            b64 = base64.b64encode(buf.getvalue()).decode("ascii")
            # Include the full payload so downstream consumers (or the model
            # via a vision-capable wrapper) can actually decode the pixels.
            # Format is a `data:` URL so clients can recognize the encoding.
            return ToolResult(f"data:image/png;base64,{b64}")
        if action == "cursor_position":
            x, y = pyautogui.position()
            return ToolResult(f"{x},{y}")
        if action == "mouse_move" and coord:
            pyautogui.moveTo(coord[0], coord[1])
            return ToolResult(f"moved to {coord[0]},{coord[1]}")
        if action == "left_click":
            if coord:
                pyautogui.click(coord[0], coord[1])
            else:
                pyautogui.click()
            return ToolResult("clicked")
        if action == "right_click":
            pyautogui.rightClick(*(coord or ()))
            return ToolResult("right-clicked")
        if action == "double_click":
            pyautogui.doubleClick(*(coord or ()))
            return ToolResult("double-clicked")
        if action == "middle_click":
            pyautogui.middleClick(*(coord or ()))
            return ToolResult("middle-clicked")
        if action == "left_click_drag" and coord:
            pyautogui.dragTo(coord[0], coord[1], duration=0.2)
            return ToolResult(f"dragged to {coord[0]},{coord[1]}")
        if action == "type" and text is not None:
            pyautogui.write(text, interval=0.01)
            return ToolResult(f"typed {len(text)} chars")
        if action == "key" and text is not None:
            # text is a key name or +-joined chord ("ctrl+c")
            keys = text.split("+")
            pyautogui.hotkey(*keys) if len(keys) > 1 else pyautogui.press(keys[0])
            return ToolResult(f"pressed {text}")
        if action == "scroll" and text is not None:
            try:
                pyautogui.scroll(int(text))
            except ValueError:
                return ToolResult("scroll text must be an integer", is_error=True)
            return ToolResult(f"scrolled {text}")
        return ToolResult(f"unsupported action: {action}", is_error=True)
    except Exception as e:  # noqa: BLE001 — surface to model
        return ToolResult(f"computer_use error: {type(e).__name__}: {e}", is_error=True)


def _plan_path(ctx: ToolContext) -> Path | None:
    root = getattr(ctx, "_plan_root", None)
    session_id = getattr(ctx, "_plan_session_id", None)
    if root is None or session_id is None:
        return None
    Path(root).mkdir(parents=True, exist_ok=True)
    safe = "".join(c for c in session_id if c.isalnum() or c in "-_") or "default"
    return Path(root) / f"{safe}.md"


def _plan_tool(args: dict[str, Any], ctx: ToolContext) -> ToolResult:
    """View / write / append / clear a persistent plan file for the session.

    The plan is a markdown scratchpad the agent can maintain across turns and
    re-read when starting work on a new turn. Based on the Manus-style
    planning pattern (github.com/OthmanAdi/planning-with-files).
    """
    path = _plan_path(ctx)
    if path is None:
        return ToolResult(
            "Plan is only available within a named session. Re-run with --session <id>.",
            is_error=True,
        )
    action = (args.get("action") or "view").lower().strip()
    if action == "view":
        if not path.exists():
            return ToolResult("(no plan yet)")
        return ToolResult(_truncate(path.read_text(encoding="utf-8")))
    if action == "write":
        content = args.get("content", "")
        path.write_text(content, encoding="utf-8")
        return ToolResult(f"Wrote {len(content)} chars to plan.")
    if action == "append":
        content = args.get("content", "")
        with path.open("a", encoding="utf-8") as f:
            f.write(content if content.endswith("\n") else content + "\n")
        return ToolResult(f"Appended {len(content)} chars to plan.")
    if action == "clear":
        if path.exists():
            path.unlink()
        return ToolResult("Plan cleared.")
    return ToolResult(f"Unknown action: {action}. Use view|write|append|clear.", is_error=True)


# ----- Builtin registry ----------------------------------------------------

def builtin_tools() -> list[Tool]:
    return [
        Tool(
            name="read_file",
            description="Read a UTF-8 text file from the working directory.",
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string", "description": "Relative path."}},
                "required": ["path"],
            },
            func=_read_file,
        ),
        Tool(
            name="list_dir",
            description="List entries in a directory under the working directory.",
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string", "default": "."}},
            },
            func=_list_dir,
        ),
        Tool(
            name="write_file",
            description="Create or overwrite a UTF-8 text file under the working directory.",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "content": {"type": "string"},
                },
                "required": ["path", "content"],
            },
            func=_write_file,
        ),
        Tool(
            name="edit_file",
            description=(
                "Replace exactly one occurrence of old_string with new_string in "
                "an existing file. Include enough surrounding context in "
                "old_string to make it unique — an ambiguous match is an error."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "old_string": {"type": "string"},
                    "new_string": {"type": "string"},
                },
                "required": ["path", "old_string"],
            },
            func=_edit_file,
        ),
        Tool(
            name="extract_doc",
            description=(
                "Extract plain text from a document. Supports .pdf, .docx, "
                ".xlsx, and any UTF-8 text file. Requires the optional `docs` "
                "extra (pip install -e 'runtime[docs]') for binary formats."
            ),
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string"}},
                "required": ["path"],
            },
            func=_extract_doc,
        ),
        Tool(
            name="run_shell",
            description=(
                "Run a shell command from a strict allowlist. Disabled unless "
                "AGENCY_ALLOW_SHELL=1 is set in the environment."
            ),
            input_schema={
                "type": "object",
                "properties": {"command": {"type": "string"}},
                "required": ["command"],
            },
            func=_run_shell,
        ),
        Tool(
            name="web_fetch",
            description="Fetch the body of an HTTP/HTTPS URL (read-only GET).",
            input_schema={
                "type": "object",
                "properties": {"url": {"type": "string"}},
                "required": ["url"],
            },
            func=_web_fetch,
        ),
        Tool(
            name="list_skills",
            description="List the other agency skills available for delegation.",
            input_schema={"type": "object", "properties": {}},
            func=_list_skills,
        ),
        Tool(
            name="delegate_to_skill",
            description=(
                "Invoke another agency skill as a sub-agent and return its final "
                "text response. Use when a task falls outside your specialty and "
                "another skill is a better fit. Look up valid slugs with list_skills."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "slug": {"type": "string", "description": "Target skill slug."},
                    "request": {"type": "string", "description": "Task to hand off."},
                },
                "required": ["slug", "request"],
            },
            func=_delegate_to_skill,
        ),
        Tool(
            name="computer_use",
            description=(
                "Control the desktop: mouse, keyboard, screenshots. Off unless "
                "AGENCY_ENABLE_COMPUTER_USE=1 AND the `[computer]` extra is "
                "installed AND a display is available. Actions: screenshot, "
                "cursor_position, mouse_move, left_click, right_click, "
                "middle_click, double_click, left_click_drag, type, key, scroll."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "action": {"type": "string"},
                    "coordinate": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "minItems": 2,
                        "maxItems": 2,
                        "description": "[x, y] for move/click/drag actions.",
                    },
                    "text": {"type": "string", "description": "For type / key / scroll."},
                },
                "required": ["action"],
            },
            func=_computer_use,
        ),
        Tool(
            name="plan",
            description=(
                "Read or update a persistent per-session markdown plan. Use it to "
                "decompose a task, track progress across turns, and re-read what "
                "you've committed to before acting. Actions: view | write | append "
                "| clear. Only available when a session id is set."
            ),
            input_schema={
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["view", "write", "append", "clear"],
                    },
                    "content": {
                        "type": "string",
                        "description": "Markdown content for write/append.",
                    },
                },
                "required": ["action"],
            },
            func=_plan_tool,
        ),
    ]


def tools_by_name(tools: list[Tool]) -> dict[str, Tool]:
    return {t.name: t for t in tools}
